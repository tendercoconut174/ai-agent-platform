"""Delivery Service – formats results for the requested output format."""

import base64
import io
import logging
import uuid
from typing import Any, Optional

from shared.models.schemas import MessageResponse

logger = logging.getLogger(__name__)

# Formats that produce file downloads (derived from available formatters)
SUPPORTED_FILE_FORMATS = frozenset({"pdf", "xl", "audio"})


def _wrap_line(text: str, width: int = 90) -> list[str]:
    words = text.split()
    lines, current, current_len = [], [], 0
    for w in words:
        if current_len + len(w) + 1 <= width:
            current.append(w)
            current_len += len(w) + 1
        else:
            if current:
                lines.append(" ".join(current))
            current, current_len = [w], len(w)
    if current:
        lines.append(" ".join(current))
    return lines


def _text_to_pdf(content: str) -> bytes:
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(True, margin=15)
    pdf.set_margins(left=15, top=15, right=15)
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    w = pdf.epw
    for block in content.split("\n"):
        for line in _wrap_line(block):
            pdf.multi_cell(w=w, h=6, txt=line)
    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


def _text_to_excel(content: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    lines = [l.strip() for l in content.split("\n") if l.strip()]
    rows = []
    for line in lines:
        if "|" in line:
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if cells and not all(c.replace("-", "").replace(":", "").strip() == "" for c in cells):
                rows.append(cells)
        else:
            rows.append([line])
    if not rows:
        ws["A1"] = content[:50000]
    else:
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def deliver(
    result: Optional[dict[str, Any]],
    workflow_id: Optional[str] = None,
    output_format: str = "json",
) -> dict[str, Any]:
    """Format result for delivery.

    Args:
        result: Dict with 'result' key containing the text output.
        workflow_id: Workflow identifier.
        output_format: json, pdf, xl, audio.

    Returns:
        Serialized MessageResponse dict.
    """
    logger.info("[delivery] deliver | workflow_id=%s | format=%s", workflow_id, output_format)
    if result is None:
        logger.warning("[delivery] result is None (timeout)")
        return MessageResponse(result="Request timed out", output_format=output_format).model_dump()

    text = result.get("result", "")

    if output_format == "pdf":
        try:
            pdf_bytes = _text_to_pdf(text)
            return MessageResponse(
                result=text,
                workflow_id=workflow_id,
                output_format="pdf",
                content_base64=base64.b64encode(pdf_bytes).decode("ascii"),
                content_type="application/pdf",
                filename=f"result_{workflow_id or uuid.uuid4()}.pdf",
            ).model_dump()
        except Exception as e:
            logger.warning("[delivery] PDF conversion failed: %s", e)
            return MessageResponse(result=text, workflow_id=workflow_id, output_format="json").model_dump()

    if output_format == "xl":
        try:
            xl_bytes = _text_to_excel(text)
            return MessageResponse(
                result=text,
                workflow_id=workflow_id,
                output_format="xl",
                content_base64=base64.b64encode(xl_bytes).decode("ascii"),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"result_{workflow_id or uuid.uuid4()}.xlsx",
            ).model_dump()
        except Exception as e:
            logger.warning("[delivery] Excel conversion failed: %s", e)
            return MessageResponse(result=text, workflow_id=workflow_id, output_format="json").model_dump()

    if output_format == "audio":
        try:
            from services.delivery.formatters.audio import text_to_audio_base64

            audio_data = text_to_audio_base64(text)
            if audio_data:
                return MessageResponse(
                    result=text,
                    workflow_id=workflow_id,
                    output_format="audio",
                    content_base64=audio_data["content_base64"],
                    content_type=audio_data["content_type"],
                    filename=audio_data["filename"],
                ).model_dump()
        except Exception as e:
            logger.warning("[delivery] Audio conversion failed: %s", e)

    return MessageResponse(result=text, workflow_id=workflow_id, output_format=output_format).model_dump()
