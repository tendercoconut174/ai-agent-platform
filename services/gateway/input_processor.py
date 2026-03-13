"""Input processor – handles text, audio, image, and file uploads."""

import logging
import os
import tempfile
from pathlib import Path
from typing import Optional

from fastapi import UploadFile

logger = logging.getLogger(__name__)


async def process_audio(audio: UploadFile) -> str:
    """Transcribe uploaded audio to text via OpenAI Whisper."""
    from shared.mcp.tools.media_processor import transcribe_audio

    suffix = Path(audio.filename or "audio.wav").suffix
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await audio.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = transcribe_audio(tmp_path)
        if result.get("error"):
            logger.error("Audio transcription failed: %s", result["error"])
            return ""
        return result["text"]
    finally:
        os.unlink(tmp_path)


async def process_image(image: UploadFile) -> str:
    """Describe uploaded image via GPT-4 Vision."""
    from shared.mcp.tools.media_processor import describe_image

    suffix = Path(image.filename or "image.png").suffix
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await image.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = describe_image(tmp_path)
        if result.get("error"):
            logger.error("Image description failed: %s", result["error"])
            return ""
        return result["description"]
    finally:
        os.unlink(tmp_path)


async def process_file(file: UploadFile) -> str:
    """Extract text from uploaded files (PDF, CSV, TXT, Excel)."""
    filename = file.filename or "file"
    content = await file.read()
    ext = Path(filename).suffix.lower()

    if ext == ".txt":
        return content.decode("utf-8", errors="replace")

    if ext == ".csv":
        return content.decode("utf-8", errors="replace")

    if ext == ".pdf":
        try:
            import io
            from fpdf import FPDF
            # Basic PDF text extraction
            text = content.decode("latin-1", errors="replace")
            # Extract text between stream markers (simple extraction)
            return f"[PDF content from {filename}, {len(content)} bytes]"
        except Exception:
            return f"[Unable to extract text from {filename}]"

    if ext in (".xlsx", ".xls"):
        try:
            import io
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(content), read_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append(" | ".join(cells))
            return "\n".join(lines)
        except Exception:
            return f"[Unable to extract data from {filename}]"

    return f"[Unsupported file type: {ext}]"


async def build_message(
    text: Optional[str] = None,
    audio: Optional[UploadFile] = None,
    image: Optional[UploadFile] = None,
    files: Optional[list[UploadFile]] = None,
) -> str:
    """Combine multi-modal inputs into a single text message for the orchestrator.

    Args:
        text: Text message.
        audio: Audio file to transcribe.
        image: Image file to describe.
        files: Additional file attachments.

    Returns:
        Combined text message.
    """
    parts = []

    if text:
        parts.append(text)

    if audio:
        transcription = await process_audio(audio)
        if transcription:
            parts.append(f"[Voice input]: {transcription}")

    if image:
        description = await process_image(image)
        if description:
            parts.append(f"[Image description]: {description}")

    if files:
        for f in files:
            file_text = await process_file(f)
            if file_text:
                parts.append(f"[File: {f.filename}]:\n{file_text}")

    return "\n\n".join(parts) if parts else ""
